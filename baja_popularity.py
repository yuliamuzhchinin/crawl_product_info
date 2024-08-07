from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
import time


def get_product_links(driver, url):
    driver.get(url)
    wait = WebDriverWait(driver, 10)
    product_links = []

    try:
        product_divs = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div.w-full')))
        for div in product_divs:
            try:
                link = div.find_element(By.CSS_SELECTOR, 'figure a[itemprop="url"]').get_attribute('href')
                if link:
                    product_links.append(link)
            except Exception as e:
                continue
    except Exception as e:
        print(f"Failed to load page {url}: {e}")

    return product_links


def get_sku(driver, url):
    driver.get(url)

    # Try to find SKU in the primary location without waiting too long
    try:
        sku_div = driver.find_element(By.CSS_SELECTOR, 'div.product-detail__sku span')
        return sku_div.text.strip().split(': ')[-1]
    except Exception:
        pass

    # If SKU is not found in the primary location, check the alternative location
    try:
        alternative_sku_div = driver.find_element(By.CSS_SELECTOR, 'div[data-show-on-vsk-sku]')
        alternative_sku_text = alternative_sku_div.get_attribute('data-show-on-vsk-sku')
        return alternative_sku_text
    except Exception:
        pass

    # If SKU is still not found, look for any "SKU:" text and grab the adjacent span values
    try:
        sku_elements = driver.find_elements(By.XPATH, "//strong[contains(text(), 'SKU:')]/span")
        if sku_elements:
            sku_values = [elem.text.strip() for elem in sku_elements]
            return '/'.join(sku_values)
    except Exception as e:
        print(f"Failed to retrieve SKU from {url}: {e}")
        return None


all_skus = []

# Set up Selenium WebDriver (use appropriate WebDriver for your browser, e.g., ChromeDriver)
options = webdriver.ChromeOptions()
options.add_argument('--headless')  # Run in headless mode
driver = webdriver.Chrome(options=options)

# Process all 44 pages
for page in range(1, 45):
    url = f"https://www.bajadesigns.com/products/?page={page}"
    print(f"\nProcessing page {page}")
    product_links = get_product_links(driver, url)

    for link in product_links:
        sku = get_sku(driver, link)
        if sku:
            all_skus.append(sku)
        print(f"Processed link: {link}, SKU: {sku}")
        time.sleep(2)  # Wait for 2 seconds between requests

driver.quit()

# Save to Excel using openpyxl
wb = Workbook()
ws = wb.active
ws.title = "SKUs"

# Add header
ws['A1'] = 'SKU'

# Add SKUs to the worksheet
for row, sku in enumerate(all_skus, start=2):
    ws[f'A{row}'] = sku

# Save the workbook
wb.save('baja_designs_skus.xlsx')

print(f"\nCollected {len(all_skus)} SKUs and saved to baja_designs_skus.xlsx")
